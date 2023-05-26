import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Protection
from pathlib import Path

print("Process has started")

output_path = "C:\\Users\\pedro\\Documents\\Repositorios\\pythonStudies\\FileManagement\\archives\\XLS MARU MATRIZ\\"
path = "C:\\Users\\pedro\\Documents\\Repositorios\\pythonStudies\\FileManagement\\archives\\02 2017-2022.xlsx"
xlsx_archive = pd.ExcelFile(path)
df = pd.read_excel(xlsx_archive)

for year in range(2017, 2024):
    for month in range(1, 13):
        year_str = str(year)
        month_str = str(month).zfill(2)
        crit1 = df['dta_cupom'].map(lambda x: x.year == year)
        crit2 = df['dta_cupom'].map(lambda x: x.month == month)
        df_filtered = df.loc[crit1 & crit2]
        df_filtered['dta_cupom'] = df_filtered['dta_cupom'].dt.strftime('%d/%m/%Y')

        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Convert the filtered DataFrame to a list of lists
        data = [list(row) for row in df_filtered.itertuples(index=False, name=None)]

        # Apply cell protection settings
        for row in data:
            ws.append(row)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.protection = Protection(locked=False, hidden=False)

        # Save the workbook with unprotected cells
        filename = month_str + '-' + year_str + ".xlsx"
        wb.save(Path(output_path) / filename)

print("Process has completed")
